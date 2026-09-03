"""Build the HIAS course data from the autumn timetable workbook."""
import json
import re
from collections import Counter, defaultdict

import openpyxl

NEW_FILE = '../2026年秋季学期课表.xlsx'
HTML_FILE = '../index.html'
OUT = 'courses_merged.json'

# ---------- 1. load existing COURSES from HTML ----------
with open(HTML_FILE, encoding='utf-8') as f:
    html = f.read()
m = re.search(r'const COURSES=(\[.*?\]);', html, re.DOTALL)
if not m:
    raise RuntimeError(f'COURSES data not found in {HTML_FILE}')
existing = json.loads(m.group(1))

# ---------- 2. load new timetable ----------
wb = openpyxl.load_workbook(NEW_FILE, data_only=True)
ws = wb['sheet0']
all_rows = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(row=r, column=col).value for col in range(1, 25)]
    if all(v is None for v in vals):
        continue
    all_rows.append(vals)

courses = []   # anchor + slots
cur = None
for vals in all_rows:
    # Course codes are the reliable anchors in the timetable workbook.
    if vals[1] is not None:
        code = str(vals[1]).strip()
        if cur is not None and code == str(cur['anchor'][1]).strip():
            # The same course can occupy multiple rows for different class times.
            cur['slots'].append(vals)
            continue
        if cur: courses.append(cur)
        cur = {'anchor': vals, 'slots': []}
    elif cur is not None:
        cur['slots'].append(vals)
if cur: courses.append(cur)

# ---------- 3. subjectCode map from existing ----------
sc_map = defaultdict(lambda: defaultdict(Counter))   # sc -> first -> Counter(second)
for c in existing:
    sc_map[c['subjectCode']][c['first']][c['second']] += 1

SPECIAL_SC = {
    '025100': ('应用经济学', '金融学'),
    '035400': ('其他 / 自设学科', '知识产权'),
    '050101': ('中国语言文学', '文艺学'),
    '050102': ('中国语言文学', '语言学及应用语言学'),
    '075100': ('大气科学', '气象学'),
    '125100': ('工商管理', '工商管理'),
    '125200': ('公共管理', '公共管理'),
    '99J100': ('其他 / 自设学科', '人居科学'),
    '99J1X1': ('其他 / 自设学科', '人居前沿及交叉科学'),
    '99J1X2': ('其他 / 自设学科', '广义建筑学'),
}

def derive_first_second(code, sc, disc):
    """Return (first, second) using subjectCode map, falling back to H discipline."""
    if sc in SPECIAL_SC:
        return SPECIAL_SC[sc]
    fs = sc_map.get(sc)
    if fs:
        firsts = sorted(fs.keys(), key=lambda k: -sum(fs[k].values()))
        first = firsts[0]
        seconds = fs[first]
        if disc and disc in seconds:
            return (first, disc)
        if disc and disc == first:
            return (first, '一级学科课程')
        second = max(seconds.items(), key=lambda kv: kv[1])[0]
        return (first, second)
    # fallback: try disc as first
    if disc:
        return (disc, disc if disc in {c['second'] for c in existing} else '一级学科课程')
    return ('其他 / 自设学科', '自设课程')

def parse_slot_m(m):
    """Parse '周二(5-6)' or '周六(1-3,5-7)' -> list of (day, start, end, label)."""
    mm = re.match(r'周([一二三四五六日天])\((.+)\)', m.strip())
    if not mm: return []
    day = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'日':7,'天':7}[mm.group(1)]
    out = []
    for part in mm.group(2).split(','):
        part = part.strip()
        if '-' in part:
            a, b = part.split('-'); out.append((day, int(a), int(b)))
        else:
            out.append((day, int(part), int(part)))
    return out

DAY_CN = {1:'周一',2:'周二',3:'周三',4:'周四',5:'周五',6:'周六',7:'周日'}

# ---------- 4. build autumn records ----------
autumn = []
for i, c in enumerate(courses, 1):
    a = c['anchor']
    college = str(a[0]).strip()
    code = str(a[1]).strip()
    name = str(a[2]).strip()
    disc = str(a[6]).strip() if a[6] else ''
    hours_credits = str(a[7]).strip().split('/')
    hours = float(hours_credits[0])
    credits = float(hours_credits[1])
    category = str(a[4]).strip()
    level = str(a[5]).strip()
    cap = a[8]
    enrolled = a[9] if a[9] is not None else 0
    exam = str(a[14]).strip() if a[14] else ''
    teach = str(a[13]).strip() if a[13] else ''
    room = str(a[12]).strip() if a[12] else ''
    english = str(a[3]).strip() if a[3] else ''

    sc = code[6:12] if len(code) >= 12 else ''
    first, second = derive_first_second(code, sc, disc)

    # campus: from code, cross-check existing
    campus = '杭高院'

    # slots
    slots = []
    slot_rows = [a] + c['slots']
    for sr in slot_rows:
        mval = sr[11]
        wval = sr[10]
        if not mval: continue
        wtext = str(wval).strip() if wval else ''
        for (day, s, e) in parse_slot_m(str(mval)):
            slots.append({'day': day, 'p': f'{s}-{e}', 'start': s, 'end': e, 'w': wtext})

    rec = {
        'id': i,
        'code': code, 'name': name, 'en': english,
        'college': college, 'campus': campus, 'semester': '秋季',
        'category': category, 'discipline': disc, 'subjectCode': sc,
        'first': first, 'second': second,
        'level': level, 'hours': hours, 'credits': credits,
        'capacity': cap, 'enrolled': int(enrolled) if enrolled else 0,
        'exam': exam, 'teach': teach,
        'chief': str(a[15]).strip() if a[15] else '', 'chiefUnit': str(a[16]).strip() if a[16] else '',
        'main': str(a[17]).strip() if a[17] else '', 'mainUnit': str(a[18]).strip() if a[18] else '',
        'ta': str(a[19]).strip() if a[19] else '', 'taUnit': str(a[20]).strip() if a[20] else '',
        'convener': str(a[21]).strip() if a[21] else '',
        'room': room, 'slots': slots,
    }
    autumn.append(rec)

# ---------- 5. assign ids, validate ----------
all_courses = autumn
for idx, c in enumerate(all_courses, 1):
    c['id'] = idx

# validation
issues = []
codes_a = [c['code'] for c in autumn]
dup_a = [k for k, v in Counter(codes_a).items() if v > 1]
if dup_a: issues.append(f'duplicate autumn codes: {dup_a[:10]}')
no_slots = [c['code'] for c in autumn if not c['slots']]
if no_slots: issues.append(f'autumn courses with no slots: {len(no_slots)} {no_slots[:5]}')
no_room = [c['code'] for c in autumn if not c['room']]
if no_room: issues.append(f'autumn courses with no room: {len(no_room)} {no_room[:5]}')
over = [c['code'] for c in autumn if c['capacity'] and c['enrolled'] > c['capacity']]
if over: issues.append(f'enrolled>capacity: {len(over)} {over[:5]}')
print('issues:', issues if issues else 'NONE')
print(f'total: autumn={len(autumn)} all={len(all_courses)}')

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(all_courses, f, ensure_ascii=False, separators=(',', ':'))
print('written', OUT)

# report stats
print('colleges union:', len({c['college'] for c in all_courses}))
print('categories:', len({c['category'] for c in all_courses}))
print('firsts:', len({c['first'] for c in all_courses}))
print('campuses:', dict(Counter(c['campus'] for c in all_courses)))
# autumn slots stats
slot_counts = Counter(len(c['slots']) for c in autumn)
print('autumn slots/course:', dict(sorted(slot_counts.items())))
